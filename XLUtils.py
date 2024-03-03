import openpyxl

def getRowCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

def readData(path, sheetName, rownum, colnum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.cell(row = rownum, column = colnum).value )

def writeData(path, sheetName, rownum, colnum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row = rownum, column = colnum).value=data
    workbook.save(path)

