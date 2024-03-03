import openpyxl

path="/Users/deepintent/Downloads/Book1.xlsx"

workbook=openpyxl.load_workbook(path)
sheet = workbook.active       #workbook.get_sheet_by_name("Sheet1")
rows = sheet.max_row
cols = sheet.max_column

print(f"Number of Rows: {rows}  \nNumber of Columns: {cols}\n")

for r in range(2,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end=" ")
    print("")